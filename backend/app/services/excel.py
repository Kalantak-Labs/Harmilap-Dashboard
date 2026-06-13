"""
Excel ingestion and export for the companies table.

A row is keyed by ISIN Code when present, otherwise by ARN Number. At least one
of the two columns must exist in the file, and each row must carry at least one.

Import column headers (exact match required):
  Company Name, ISIN Code, ARN Number, RTA Code,
  Email 1, Email 2, ... (any number of email columns)
  Contact Number 1, Contact Number 2, ... (any number of phone columns)
  Authorized Person name, Designation of Authorized Person,
  GST number, TAN number, PAN number,
  Address Line 1, Address Line 2, Address Line 3, Address Line 4,
  City, Pin Code, Billing Address, Security Type,
  Total Shares, Has NSDL Shares?, NSDL Shares,
  Has CDSL Shares?, CDSL Shares, Physical Shares
"""

import io
import re
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Mapping: Excel column header -> model field name (excludes email/phone — handled dynamically)
COLUMN_MAP: dict[str, str] = {
    "Company Name": "company_name",
    "ISIN Code": "isin_code",
    "ARN Number": "arn_number",
    "ARN": "arn_number",
    "RTA Code": "nsdl_rta_code",
    "NSDL RTA Code": "nsdl_rta_code",
    "CDSL RTA Code": "cdsl_rta_code",
    "Authorized Person name": "authorized_person_name",
    "Designation of Authorized Person": "authorized_person_designation",
    "GST number": "gst_number",
    "TAN number": "tan_number",
    "PAN number": "pan_number",
    "Address Line 1": "reg_address_line1",
    "Address Line 2": "reg_address_line2",
    "Address Line 3": "reg_address_line3",
    "Address Line 4": "reg_address_line4",
    "City": "reg_city",
    "Pin Code": "reg_pin_code",
    "Billing Address": "billing_address",
    "Security Type": "security_type",
    "Total Shares": "total_shares",
    "Has NSDL Shares?": "has_nsdl_shares",
    "NSDL Shares": "nsdl_shares",
    "Has CDSL Shares?": "has_cdsl_shares",
    "CDSL Shares": "cdsl_shares",
    "Physical Shares": "physical_shares",
}

BOOL_FIELDS = {"has_nsdl_shares", "has_cdsl_shares"}
INT_FIELDS = {"total_shares", "nsdl_shares", "cdsl_shares", "physical_shares"}

# Columns before email/phone in export
EXPORT_COLUMNS_PRE = [
    "company_name", "isin_code", "arn_number", "nsdl_rta_code", "cdsl_rta_code",
]
# Columns after email/phone in export
EXPORT_COLUMNS_POST = [
    "authorized_person_name", "authorized_person_designation",
    "gst_number", "tan_number", "pan_number",
    "reg_address_line1", "reg_address_line2", "reg_address_line3", "reg_address_line4",
    "reg_city", "reg_pin_code", "billing_address", "security_type",
    "total_shares", "has_nsdl_shares", "nsdl_shares",
    "has_cdsl_shares", "cdsl_shares", "physical_shares",
]

EXPORT_HEADERS_PRE = ["Company Name", "ISIN Code", "ARN Number", "NSDL RTA Code", "CDSL RTA Code"]
EXPORT_HEADERS_POST = [
    "Authorized Person name", "Designation of Authorized Person",
    "GST number", "TAN number", "PAN number",
    "Address Line 1", "Address Line 2", "Address Line 3", "Address Line 4",
    "City", "Pin Code", "Billing Address", "Security Type",
    "Total Shares", "Has NSDL Shares?", "NSDL Shares",
    "Has CDSL Shares?", "CDSL Shares", "Physical Shares",
]


def _sorted_numbered_cols(df_columns: list[str], prefix: str) -> list[str]:
    """Return columns matching '<prefix> <N>' sorted by N."""
    pat = re.compile(r'^' + re.escape(prefix) + r'\s+(\d+)$', re.IGNORECASE)
    matches = [(int(m.group(1)), col) for col in df_columns if (m := pat.match(col))]
    return [col for _, col in sorted(matches)]


def _clean_str(val: Any) -> str | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    return str(val).strip() or None


def _clean_bool(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() in {"true", "yes", "1", "y"}
    return False


def _clean_int(val: Any) -> int | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _clean_array(val: Any) -> list[str]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return []
    return [v.strip() for v in str(val).split(",") if v.strip()]


def parse_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Parse uploaded Excel file.
    Returns (rows, errors) where each row is a dict of model fields.

    Email columns:        Email 1, Email 2, … → email_ids (array)
    Phone columns:        Contact Number 1, Contact Number 2, … → contact_numbers (array)
    All other columns map via COLUMN_MAP.
    """
    errors: list[str] = []
    rows: list[dict] = []

    try:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    except Exception as e:
        return [], [f"Could not read Excel file: {e}"]

    all_cols = list(df.columns)
    known_cols = {col: COLUMN_MAP[col] for col in all_cols if col in COLUMN_MAP}
    email_cols = _sorted_numbered_cols(all_cols, "Email")
    phone_cols = _sorted_numbered_cols(all_cols, "Contact Number")

    has_isin_col = "ISIN Code" in known_cols
    has_arn_col = any(col in known_cols and known_cols[col] == "arn_number" for col in known_cols)
    if not has_isin_col and not has_arn_col:
        return [], ["The Excel file must contain an 'ISIN Code' or 'ARN Number' column."]

    for idx, row in df.iterrows():
        row_num = idx + 2
        isin_raw = _clean_str(row.get("ISIN Code")) if has_isin_col else None
        arn_raw = None
        for col, field in known_cols.items():
            if field == "arn_number" and (val := _clean_str(row.get(col))):
                arn_raw = val
                break
        if not isin_raw and not arn_raw:
            errors.append(f"Row {row_num}: neither ISIN Code nor ARN Number present, skipping.")
            continue

        record: dict = {}
        for excel_col, field in known_cols.items():
            raw = row.get(excel_col)
            if field in BOOL_FIELDS:
                record[field] = _clean_bool(raw)
            elif field in INT_FIELDS:
                record[field] = _clean_int(raw)
            else:
                record[field] = _clean_str(raw)

        # Collect numbered email/phone columns into arrays
        record["email_ids"] = [v for col in email_cols if (v := _clean_str(row.get(col)))]
        record["contact_numbers"] = [v for col in phone_cols if (v := _clean_str(row.get(col)))]

        rows.append(record)

    return rows, errors


def build_export_excel(companies: list[dict]) -> bytes:
    """Build a styled Excel file. Email/phone arrays expand to numbered columns."""
    max_emails = max((len(c.get("email_ids") or []) for c in companies), default=1)
    max_phones = max((len(c.get("contact_numbers") or []) for c in companies), default=1)
    max_emails = max(max_emails, 1)
    max_phones = max(max_phones, 1)

    email_headers = [f"Email {i}" for i in range(1, max_emails + 1)]
    phone_headers = [f"Contact Number {i}" for i in range(1, max_phones + 1)]

    all_headers = EXPORT_HEADERS_PRE + email_headers + phone_headers + EXPORT_HEADERS_POST
    # Build (field_or_special, index_or_None) column spec
    col_specs: list[tuple[str, int | None]] = (
        [(f, None) for f in EXPORT_COLUMNS_PRE]
        + [("email_ids", i) for i in range(max_emails)]
        + [("contact_numbers", i) for i in range(max_phones)]
        + [(f, None) for f in EXPORT_COLUMNS_POST]
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"

    header_fill = PatternFill("solid", fgColor="1a1a1a")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 16)

    ws.row_dimensions[1].height = 22

    for row_idx, company in enumerate(companies, 2):
        for col_idx, (field, arr_idx) in enumerate(col_specs, 1):
            val = company.get(field)
            if arr_idx is not None:
                arr = val if isinstance(val, list) else []
                val = arr[arr_idx] if arr_idx < len(arr) else None
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            ws.cell(row=row_idx, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


BENEF_EXPORT_HEADERS = [
    "ISIN Code", "DP ID", "Client ID", "Record Date",
    "First Holder Name", "First Holder PAN", "First Holder Email",
    "Second Holder Name", "Second Holder PAN",
    "Third Holder Name", "Third Holder PAN",
    "Beneficiary Type", "Account Category", "Status",
    "Address Line 1", "Address Line 2", "Address Line 3", "Address Line 4", "Pin Code",
    "Bank Account Number", "Bank Name & Branch", "IFSC", "MICR Code", "Bank Account Type",
    "Free Positions", "Lock-in Positions", "Block Positions", "Pledged Positions",
    "Remat Positions", "IDD Positions", "CM Pool Positions", "CC Settlement Positions",
    "Minor", "RGESS Flag",
]

BENEF_EXPORT_FIELDS = [
    "isin_code", "dp_id", "client_id", "record_date",
    "first_holder_name", "first_holder_pan", "first_holder_email",
    "second_holder_name", "second_holder_pan",
    "third_holder_name", "third_holder_pan",
    "beneficiary_type", "account_category", "beneficiary_status",
    "address_line1", "address_line2", "address_line3", "address_line4", "pin_code",
    "bank_account_number", "bank_name_branch", "ifsc", "micr_code", "bank_account_type",
    "free_positions", "lockin_positions", "block_positions", "pledged_positions",
    "remat_positions", "idd_positions", "cm_pool_positions", "cc_settlement_positions",
    "minor_indicator", "rgess_flag",
]


def build_beneficiary_export(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Beneficiaries"

    header_fill = PatternFill("solid", fgColor="1a1a1a")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(BENEF_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 14)

    ws.row_dimensions[1].height = 22

    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(BENEF_EXPORT_FIELDS, 1):
            val = row.get(field)
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

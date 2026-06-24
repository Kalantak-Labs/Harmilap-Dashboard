"""
Excel export / import for company-level tax invoices.

The sheet has a fixed set of key + payment columns, plus one dynamic column per
particular (column header = the particular's description). On import, a row is
keyed by its RTA code(s); particular columns are matched to the template by
description, and amounts are upserted onto the invoice's particulars.

Fixed columns:
  NSDL RTA Code, CDSL RTA Code, Company Name, Invoice No,
  <Particular 1>, <Particular 2>, ... (dynamic),
  Payment Status, Payment Date, Amount Paid
"""

import io
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FIXED_LEAD = ["NSDL RTA Code", "CDSL RTA Code", "Company Name", "Invoice No"]
FIXED_TAIL = ["Payment Status", "Payment Date", "Amount Paid"]


def _clean_str(val: Any) -> str | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    return str(val).strip() or None


def _clean_bool(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() in {"true", "yes", "1", "y", "paid"}
    return False


def _clean_float(val: Any) -> float | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _clean_date(val: Any) -> date | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def build_invoice_export(rows: list[dict], particular_names: list[str]) -> bytes:
    """
    rows: list of dicts with keys
      nsdl_rta_code, cdsl_rta_code, company_name, invoice_no,
      particular_amounts (dict[name -> amount]),
      payment_status (bool), payment_date (date|None), amount_paid (float|None)
    particular_names: ordered list of particular descriptions → dynamic columns
    """
    headers = FIXED_LEAD + list(particular_names) + FIXED_TAIL

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    header_fill = PatternFill("solid", fgColor="1a1a1a")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 16)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        amounts = row.get("particular_amounts") or {}
        pay_date = row.get("payment_date")
        values = (
            [row.get("nsdl_rta_code"), row.get("cdsl_rta_code"),
             row.get("company_name"), row.get("invoice_no")]
            + [amounts.get(name) for name in particular_names]
            + ["Yes" if row.get("payment_status") else "No",
               pay_date.strftime("%d.%m.%Y") if isinstance(pay_date, (date, datetime)) else None,
               row.get("amount_paid")]
        )
        for c_idx, v in enumerate(values, 1):
            ws.cell(row=r_idx, column=c_idx, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_invoice_excel(file_bytes: bytes, particular_names: list[str]) -> tuple[list[dict], list[str]]:
    """
    Returns (rows, errors). Each row:
      {nsdl_rta_code, cdsl_rta_code, particular_amounts: {name: amount},
       payment_status, payment_date, amount_paid}
    A row needs at least one RTA code. Particular columns are matched by header
    against `particular_names`; unknown particular columns are ignored.
    """
    errors: list[str] = []
    rows: list[dict] = []
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    except Exception as e:
        return [], [f"Could not read Excel file: {e}"]

    cols = list(df.columns)
    if "NSDL RTA Code" not in cols and "CDSL RTA Code" not in cols:
        return [], ["The Excel file must contain a 'NSDL RTA Code' or 'CDSL RTA Code' column."]

    name_set = {n for n in particular_names}
    present_particulars = [c for c in cols if c in name_set]

    for idx, row in df.iterrows():
        row_num = idx + 2
        nsdl = _clean_str(row.get("NSDL RTA Code"))
        cdsl = _clean_str(row.get("CDSL RTA Code"))
        if not nsdl and not cdsl:
            errors.append(f"Row {row_num}: no RTA code present, skipping.")
            continue
        amounts = {name: _clean_float(row.get(name)) for name in present_particulars}
        rows.append({
            "nsdl_rta_code": nsdl,
            "cdsl_rta_code": cdsl,
            "particular_amounts": {k: v for k, v in amounts.items() if v is not None},
            "payment_status": _clean_bool(row.get("Payment Status")),
            "payment_date": _clean_date(row.get("Payment Date")),
            "amount_paid": _clean_float(row.get("Amount Paid")),
        })

    return rows, errors


def build_billing_invoices_export(rows: list[dict]) -> bytes:
    """All billing invoices, one row each, with year-wise pending breakdown."""
    headers = [
        "Invoice No", "Company", "NSDL RTA", "CDSL RTA", "Invoice Date", "Financial Year",
        "Total Active ISINs", "Billed ISINs", "Year-wise Pending", "Grand Total (Rs.)",
        "Type", "Generated On",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    header_fill = PatternFill("solid", fgColor="1a1a1a")
    header_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = max(len(h) + 4, 16)
    ws.row_dimensions[1].height = 22

    for ri, r in enumerate(rows, 2):
        yb = r.get("year_breakdown") or []
        yb_text = ", ".join(f"FY {y.get('fiscal_year')}: {y.get('isin_count')}" for y in yb)
        gen = r.get("generated_at")
        idt = r.get("invoice_date")
        vals = [
            r.get("invoice_no"), r.get("company_name"), r.get("nsdl_rta_code"),
            r.get("cdsl_rta_code"),
            idt.strftime("%d.%m.%Y") if hasattr(idt, "strftime") else idt,
            r.get("fiscal_year"), r.get("isin_total"), r.get("billed_isin_count"),
            yb_text, r.get("grand_total"),
            "Manual" if r.get("is_manual") else "Generated",
            gen.strftime("%d.%m.%Y %H:%M") if hasattr(gen, "strftime") else gen,
        ]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
